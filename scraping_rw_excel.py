import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

chrome_options = Options()
chrome_options.add_argument("--kiosk")

chrome = webdriver.Chrome(options= chrome_options)

excel_file_name = "excel_file.xlsx"
wb = openpyxl.load_workbook(excel_file_name)

sheet = wb.get_sheet_by_name('Sheet1')

for n in range(2, sheet.max_row + 1):
    print(sheet.cell(row=n, column=2).value)
    chrome.get(sheet.cell(row=n, column=2).value)
    sheet.cell(row=n, column=3).value = 'no'

wb.save(excel_file_name)